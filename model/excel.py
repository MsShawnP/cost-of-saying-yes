"""CFO-grade Excel workbook builder.

Produces a 4-tab workbook:
  - Summary: key metrics across all three scenarios
  - Realistic, Optimistic, Pessimistic: full 12-month detail per scenario

Design rules:
  - Register ALL NamedStyles on the workbook BEFORE writing any cells (openpyxl
    raises KeyError if you reference a style that hasn't been registered yet)
  - wb.remove(wb.active) immediately after Workbook() to drop the default Sheet1
  - BytesIO streaming — no disk writes
  - freeze_panes = "A2" on every sheet (freezes header row only)
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Lailara colors in aRGB (openpyxl expects AARRGGBB hex without #)
NAVY       = "FF1F2E7A"   # Chicago navy
WHITE      = "FFFFFFFF"
CANVAS     = "FFF5F3EE"   # Lailara canvas
LIGHT_GRAY = "FFE8E8E8"


def _register_styles(wb: Workbook) -> None:
    """Register all NamedStyles BEFORE any cell writes."""

    # Header style — navy background, white bold text
    header = NamedStyle(name="header")
    header.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    header.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    header.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(header)

    # Sub-header — light gray background
    sub_header = NamedStyle(name="sub_header")
    sub_header.font = Font(name="Calibri", bold=True, size=10)
    sub_header.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)
    sub_header.alignment = Alignment(horizontal="center")
    wb.add_named_style(sub_header)

    # Currency — $#,##0
    currency = NamedStyle(name="currency")
    currency.font = Font(name="Calibri", size=10)
    currency.number_format = '"$"#,##0'
    currency.alignment = Alignment(horizontal="right")
    wb.add_named_style(currency)

    # Negative currency — red parentheses for negative values
    neg_currency = NamedStyle(name="neg_currency")
    neg_currency.font = Font(name="Calibri", size=10)
    neg_currency.number_format = '"$"#,##0;[Red]("$"#,##0)'
    neg_currency.alignment = Alignment(horizontal="right")
    wb.add_named_style(neg_currency)

    # Percentage
    pct = NamedStyle(name="pct")
    pct.font = Font(name="Calibri", size=10)
    pct.number_format = "0.0%"
    pct.alignment = Alignment(horizontal="right")
    wb.add_named_style(pct)

    # Plain text
    plain = NamedStyle(name="plain")
    plain.font = Font(name="Calibri", size=10)
    wb.add_named_style(plain)

    # Bold plain
    bold_plain = NamedStyle(name="bold_plain")
    bold_plain.font = Font(name="Calibri", bold=True, size=10)
    wb.add_named_style(bold_plain)


def _set_print_settings(ws, last_row: int, last_col: int) -> None:
    """Apply print settings — landscape, fit to 1 page wide, footer."""
    col_letter = get_column_letter(last_col)
    ws.print_area = f"A1:{col_letter}{last_row}"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.oddFooter.left.text = "Confidential — Lailara LLC"
    ws.oddFooter.right.text = "Page &P of &N"


def _build_summary_sheet(wb: Workbook, scenarios: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 34
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 20

    # Header row
    ws.row_dimensions[1].height = 20
    ws["A1"].value = "Metric"
    ws["A1"].style = "header"
    ws["B1"].value = "Realistic"
    ws["B1"].style = "header"
    ws["C1"].value = "Optimistic"
    ws["C1"].style = "header"
    ws["D1"].value = "Pessimistic"
    ws["D1"].style = "header"

    # Row definitions: (label, summary_key, style)
    rows = [
        ("Gross Revenue — Year 1",     "gross_revenue_year1",    "currency"),
        ("Total Deductions — Year 1",  "total_deductions_year1", "neg_currency"),
        ("Net Revenue — Year 1",       "net_revenue_year1",      "currency"),
        ("Upfront Investment",         "upfront_investment",     "neg_currency"),
        ("COGS — Year 1",              "cogs_year1",             "neg_currency"),
        ("Net Cash Impact — Year 1",   "net_cash_impact_year1",  "neg_currency"),
        ("Break-Even Month",           "break_even_month",       "plain"),
        ("Broker Projection Year 1",   "broker_projection_year1","currency"),
    ]

    for row_idx, (label, key, style) in enumerate(rows, start=2):
        ws.cell(row_idx, 1).value = label
        ws.cell(row_idx, 1).style = "bold_plain"
        for col_idx, scenario_key in enumerate(["realistic", "optimistic", "pessimistic"], start=2):
            val = scenarios[scenario_key]["summary"].get(key)
            cell = ws.cell(row_idx, col_idx)
            # break_even_month can be None
            if val is None:
                cell.value = "No break-even in 12 months"
                cell.style = "plain"
            else:
                cell.value = val
                cell.style = style

    _set_print_settings(ws, len(rows) + 1, 4)


def _build_scenario_sheet(wb: Workbook, scenario_name: str, scenario_data: dict) -> None:
    title = scenario_name.capitalize()
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 10   # Month
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 18

    # Header
    headers = ["Month", "Gross Revenue", "Deductions", "Net Invoiced", "Cash Received", "Cumulative Cash"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = h
        cell.style = "header"

    # Data rows
    months       = scenario_data["months"]
    gross_rev    = scenario_data["gross_revenue"]
    deductions   = scenario_data["deductions"]
    cash_recv    = scenario_data["cash_received"]
    cumulative   = scenario_data["cumulative_cash_position"]

    for m_idx, month in enumerate(months):
        row = m_idx + 2
        net_inv = gross_rev[m_idx] - deductions[m_idx]

        ws.cell(row, 1).value = month
        ws.cell(row, 1).style = "plain"

        ws.cell(row, 2).value = gross_rev[m_idx]
        ws.cell(row, 2).style = "currency"

        ws.cell(row, 3).value = -deductions[m_idx]  # negative so neg_currency renders red
        ws.cell(row, 3).style = "neg_currency"

        ws.cell(row, 4).value = net_inv
        ws.cell(row, 4).style = "currency"

        ws.cell(row, 5).value = cash_recv[m_idx]
        ws.cell(row, 5).style = "currency"

        ws.cell(row, 6).value = cumulative[m_idx]
        ws.cell(row, 6).style = "neg_currency"

    _set_print_settings(ws, len(months) + 1, 6)


def build_excel_workbook(scenarios: dict) -> Workbook:
    """Build the 4-tab CFO-grade workbook.

    Args:
        scenarios: dict with keys 'realistic', 'optimistic', 'pessimistic',
                   each value being a dict matching the API response shape.
    Returns:
        openpyxl Workbook (not yet saved — caller streams via BytesIO).
    """
    wb = Workbook()
    wb.remove(wb.active)    # drop default Sheet1 immediately

    _register_styles(wb)    # MUST happen before any cell.style = "..." assignments

    _build_summary_sheet(wb, scenarios)
    for scenario_key in ("realistic", "optimistic", "pessimistic"):
        _build_scenario_sheet(wb, scenario_key, scenarios[scenario_key])

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Save workbook to BytesIO and return bytes."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
